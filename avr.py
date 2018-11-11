import os
import openpyxl
import re
import requests
from lxml import html

pth = os.getcwd()
os.chdir(pth)


def clear_str():
    if os.name == "posix":
        os.system("clear")
    elif os.name == "nt":
        os.system("cls")

clear_str()
# Открываем экселевский файл с данными
while True:
    index = input("Индекс отделения: ")
    try:
        wb_in = openpyxl.load_workbook(index + ".xlsx")
        break
    except:
        print(f"Файла КЕ по {index} не существует \n Попробуйте снова")
wb_out = openpyxl.load_workbook("template.xlsx")
# выбираем первый лист экселевской книги
sheet_in = wb_in.worksheets[0]
sheet_out = wb_out.worksheets[0]
# в цикле по количеству строк выводим нужные значения
s_in = 1
s_out = 10
for i in range(sheet_in.max_row - 1):
    s_in += 1
    model = re.sub(r'[(].*[)]', '', sheet_in['K' + str(s_in)].value)
    serial = sheet_in['I' + str(s_in)].value
    sheet_out["B" + str(s_out)].value = i + 1
    sheet_out["C" + str(s_out)].value = model
    sheet_out["E" + str(s_out)].value = serial
    sheet_out["I" + str(s_out)].value = "Плановое техническое обслуживание"
    s_out += 1

# Ищем адрес по интексу (При наличии интернета)
try:
    response = requests.get(f'http://index-post-address.ru/address/{index}')
    tree = html.fromstring(response.text)
    address  = tree.xpath("//center/table/tr[1]/td/table/tr[6]/td[2]")[0].text_content()
    address = address.replace('\r', '').replace('\n', '').split(",")
    sheet_out["E6"].value = f"{address[2]}, {address[0]}, {address[1]}"    
except:
    sheet_out["E6"].value = f"ОПС {index}"


wb_out.save(f"new_avr/avr_to_{index}.xlsx")
clear_str()
if input("Посмотреть полученый АВР? (y/n)") == "y":
    if os.name == "posix":
        os.system(f"libreoffice --calc new_avr/avr_to_{index}.xlsx")
    elif os.name == "nt":
        os.system(f"new_avr/avr_to_{index}.xlsx")
